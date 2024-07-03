import json
import os
import openpyxl
import shutil

import requests
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook import Workbook

NAME_MONTH = {
        '1' : 'Январь',
        '2' : 'Февраль',
        '3': 'Март',
        '4': 'Апрель',
        '5': 'Май',
        '6': 'Июнь',
        '7': 'Июль',
        '8': 'Август',
        '9': 'Сентябрь',
        '10': 'Октябрь',
        '11': 'Ноябрь',
        '12': 'Декабрь',
    }
LIST_31_MONTH = [1, 3, 5, 7, 8, 10, 12]

class Date:
    'Дата при создании Таблиц'
    def __init__(self, string_date):
        date = list(map(int, string_date.split('-')))
        self.year = date[0]
        self.month = date[1]
        self.day = date[2]

    def GetDate(self):
        return(f'{self.year}-{self.month}-{self.day}')

    def IncYear(self):
        self.year += 1

    def IncMonth(self):
        self.month += 1
        if self.month == 13:
            self.month = 1
            self.IncYear()

    def IncDay(self):
        self.day += 1
        if self.day == 32:
            self.day = 1
            self.IncMonth()
        elif self.day == 31:
            if not (self.month in LIST_31_MONTH):
                self.IncMonth()
        elif self.day == 30:
            if self.month == 2:
                self.day = 1
                self.IncMonth()
        elif self.day == 29:
            if self.month == 2:
                if not (self.year % 400 == 0 or (self.year % 4 == 0 and self.year % 100 != 0)):
                    self.day = 1
                    self.IncMonth()

#region API Methods
def GetToken(_config):
    _token = ''
    try:
        _url = f'http://{_config["serverIP"]}/api/system/auth'
        _payload = {
            "login": _config['login'],
            "password": _config['password']
        }
        _headers = {
            "Content-type": "application/json; charset=UTF-8",
            #"Authorization": "Bearer rJByeM5PJanwBErcGAtIoBG5CitfQFNj"
        }
        _token = requests.request("post", _url, json=_payload, headers=_headers, timeout=0.2).json()['token']
    except Exception as e:
        print(f'Ошибка при авторизации: {e}')
    return _token

def GetAccessReport(_config, _token):

    def ResponceToString(_responce, _config):
        _answer_string = ''
        for user in responce['rows']:
            flag = True
            for word in _config['wrong_words']:
                if word in user['fio']:
                    flag = False
                    break
            if flag:
                _answer_string += f'{user["fio"]}|{user["position_name"]}|{user["division_name"]}|{user["zone_exit"]}|{user["zone_enter"]}|{user["time_label"]}\n'
        return _answer_string

    answer = ''
    try:
        _url = f'http://{_config["serverIP"]}/api/accessReports/events'
        payload = {
            'token': _token,
            'page': 1,
            'sord': 'asc',
            'dateBegin': _config["dateBegin"],
            'dateEnd': _config["dateEnd"],
            'searchString': 'КПП производства'
        }
        headers = {
            "Content-type": "application/json; charset=UTF-8",
            # "Authorization": "Bearer rJByeM5PJanwBErcGAtIoBG5CitfQFNj"
        }
        responce = requests.request('get', _url, params=payload, headers=headers).json()
        count_pages = responce['total']
        print(f'Страница {payload["page"]} из {count_pages}')
        answer += ResponceToString(responce, _config)
        if count_pages != 0:
            for i in range(count_pages - 1):
                payload['page'] += 1
                responce = requests.request('get', _url, params=payload, headers=headers).json()
                print(f'Страница {payload["page"]} из {count_pages}')
                answer += ResponceToString(responce, _config)
    except Exception as e:
        print(f'Ошибка при запросе данных о проходе')
    return answer
#endregion

def StringToArray(_accessReport):
    _array = list()
    _userList = list()
    _user = ''
    _accessArray = _accessReport.split('\n')
    for access in _accessArray:
        row = access.split('|')
        if _user == row[0]:
            _userList.append(row)
        else:
            if len(_userList) != 0:
                _userList = sorted(_userList, key=lambda x: x[5], reverse=False)
                _array.append(_userList)
            _userList = list()
            _userList.append(row)
            _user = row[0]
    return _array

def DeleteFolder(_path):
    _flag = True
    try:
        shutil.rmtree(_path)
        os.mkdir(_path)
    except Exception as e:
        print(f'Произошла ошибка при очистке директории. {e}')
        _flag = False
    return _flag

def SaveReports(_arrayAccessReports):
    _flag = True
    if DeleteFolder('Информация о проходах'):
        try:
            for _user in _arrayAccessReports:
                wb = Workbook()
                ws = wb.active
                _name = _user[0][0]
                print(_name)
                for _enter in _user:
                    row = (_enter)
                    ws.append(row)
                wb.save("Информация о проходах/" + _name + ".xlsx")
        except Exception as e:
            print('При заполнении таблицы о проходах возникла ошибка')
            _flag = False
    else:
        _flag = False
    return _flag

def NumberOfSheets(_date_begin_str, _date_end_str):
    _date_begin = Date(_date_begin_str)
    _date_end = Date(_date_end_str)
    #Step 1
    _count = 2 * (12 * (_date_end.year - _date_begin.year) + _date_end.month - _date_begin.month - 1)
    if _date_begin.day <= 15:
        _count += 2
    else:
        _count += 1
    if _date_end.day <= 15:
        _count += 1
    else:
        _count += 2
    return _count

def CreateTemplateHeader(_worksheet, _date, _half_month):

    def GetRightBorder(_date, _half_month):
        _right_border = 19
        if _half_month == 1:
            if _date.month in LIST_31_MONTH:
                _right_border = 20
            elif _date.month == 2:
                if _date.year % 400 == 0 or (_date.year % 4 == 0 and _date.year % 100 != 0):
                    _right_border = 18
                else:
                    _right_border = 17
        else:
            _right_border = 19
        return _right_border

    def SetFirstRow(_worksheet, _thin):
        # Первая строка заголовка
        _worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_right_border)
        _cell = _worksheet["A1"]
        for j in range(1, _right_border + 1):
            _worksheet.cell(1, j).border = Border(top=_thin, left=_thin, right=_thin, bottom=_thin)
        _cell.value = f'Табель сотрудников ООО "ПРАНАФАРМ" за {NAME_MONTH[str(_date.month)]} {_date.year} ул. Дзержинского, д.46'
        _cell.fill = PatternFill("solid", fgColor="CCCCCC")
        _cell.font = Font(name="Times New Roman", size=12, bold=True)
        _cell.alignment = Alignment(horizontal="center", vertical="center")

    def SetNumberColumn(_worksheet, _thin):
        # №
        _worksheet.column_dimensions['A'].width = 6
        _worksheet.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        _cell = _worksheet["A2"]
        for i in range(2, 4):
            _worksheet.cell(i, 1).border = Border(top=_thin, left=_thin, right=_thin, bottom=_thin)
        _cell.value = f'№'
        _cell.fill = PatternFill("solid", fgColor="CCCCCC")
        _cell.font = Font(name="Times New Roman", size=12, bold=True)
        _cell.alignment = Alignment(horizontal="center", vertical="center")

    def SetFIOColumn(_worksheet, _thin):
        # ФИО
        _worksheet.column_dimensions['B'].width = 25
        _worksheet.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
        _cell = _worksheet["B2"]
        for i in range(2, 4):
            _worksheet.cell(i, 2).border = Border(top=_thin, left=_thin, right=_thin, bottom=_thin)
        _cell.value = f'ФИО'
        _cell.fill = PatternFill("solid", fgColor="CCCCCC")
        _cell.font = Font(name="Times New Roman", size=12, bold=True)
        _cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def SetPozitionNameColumn(_worksheet, _thin):
        # Должность
        _worksheet.column_dimensions['C'].width = 25
        _worksheet.merge_cells(start_row=2, start_column=3, end_row=3, end_column=3)
        _cell = _worksheet["C2"]
        for i in range(2, 4):
            _worksheet.cell(i, 3).border = Border(top=_thin, left=_thin, right=_thin, bottom=_thin)
        _cell.value = f'Должность'
        _cell.fill = PatternFill("solid", fgColor="CCCCCC")
        _cell.font = Font(name="Times New Roman", size=12, bold=True)
        _cell.alignment = Alignment(horizontal="center", vertical="center")

    def SetDivisionNameColumn(_worksheet, _thin):
        # Подразделение
        _worksheet.column_dimensions['D'].width = 25
        _worksheet.merge_cells(start_row=2, start_column=4, end_row=3, end_column=4)
        _cell = _worksheet["D2"]
        for i in range(2, 4):
            _worksheet.cell(i, 4).border = Border(top=_thin, left=_thin, right=_thin, bottom=_thin)
        _cell.value = f'Подразделение'
        _cell.fill = PatternFill("solid", fgColor="CCCCCC")
        _cell.font = Font(name="Times New Roman", size=12, bold=True)
        _cell.alignment = Alignment(horizontal="center", vertical="center")

    def SetSecondRow(_worksheet, _thin, _right_border):
        # Вторая строка заголовка
        _thin = Side(border_style="medium", color="000000")
        _worksheet.merge_cells(start_row=2, start_column=5, end_row=2, end_column=_right_border)
        _cell = _worksheet["E2"]
        for j in range(5, _right_border + 1):
            _worksheet.cell(2, j).border = Border(top=_thin, left=_thin, right=_thin, bottom=_thin)
        _cell.value = f'{_half_month + 1}-ая половина за {NAME_MONTH[str(_date.month)]} {_date.year}'
        _cell.fill = PatternFill("solid", fgColor="CCCCCC")
        _cell.font = Font(name="Times New Roman", size=12, bold=True)
        _cell.alignment = Alignment(horizontal="center", vertical="center")

    def SetDaysRow(_worksheet, _thin, _right_border):
        _begin_day = 1
        if _half_month == 1:
            _begin_day = 16
        # Числа(дата)
        for col in range(5, _right_border + 1):
            _cell = _worksheet.cell(3, col)
            _cell.border = Border(top=_thin, left=_thin, right=_thin, bottom=_thin)
            _cell.value = _begin_day
            _cell.fill = PatternFill("solid", fgColor="CCCCCC")
            _cell.font = Font(name="Times New Roman", size=12, bold=True)
            _cell.alignment = Alignment(horizontal="center", vertical="center")
            _begin_day += 1

    _right_border = GetRightBorder(_date, _half_month)
    _thin = Side(border_style="medium", color="000000")
    SetFirstRow(_worksheet, _thin)
    SetNumberColumn(_worksheet, _thin)
    SetFIOColumn(_worksheet, _thin)
    SetPozitionNameColumn(_worksheet, _thin)
    SetDivisionNameColumn(_worksheet, _thin)
    SetSecondRow(_worksheet, _thin, _right_border)
    SetDaysRow(_worksheet, _thin, _right_border)

def CreateTabelTemplate(_count, _beginDate):
    flag = True
    try:
        _wb = Workbook()
        _date = Date(_beginDate)
        _half_month = 0
        if _date.day > 15:
            _half_month = 1
        for i in range(_count):
            _ws = _wb.create_sheet()
            _ws.title = str(i)
            CreateTemplateHeader(_ws, _date, _half_month)
            _half_month += 1
            _date.day += 1
            if _half_month == 2:
                _date.IncMonth()
                _half_month = 0
        _wb.remove(_wb['Sheet'])
        _wb.save('Табель.xlsx')
    except Exception as e:
        print(f'Во время записи шаблона табеля возникла ошибка. Данные не верны. {e}')
        flag = False
    return flag

def LoadMainTabel(date_end, count_all_sheets):
    main_tabel_book = openpyxl.load_workbook('Табель.xlsx')
    #print(main_tabel_book.sheetnames)
    for document in os.listdir('Информация о проходах'):
        user_workbook = openpyxl.load_workbook('Информация о проходах/' + document)
        user_worksheet = user_workbook.active
        user_name = user_worksheet.cell(1, 1).value
        user_position_name = user_worksheet.cell(1, 2).value
        user_division_name = user_worksheet.cell(1, 3).value
        data_worker = user_name + '!' + user_division_name + '!' + user_position_name
        print(data_worker)
        cur_datetime = user_worksheet.cell(1, 6).value
        cur_date = Date(cur_datetime.split()[0])
        enter_time = '00:00:00'
        exit_time = '24:00:00'
        if user_worksheet.cell(1, 4).value == 'КПП производства':
            enter_time = cur_datetime.split()[1]
        else:
            exit_time = cur_datetime.split()[1]
        for i in range (2, user_worksheet.max_row + 1):
            next_datetime = user_worksheet.cell(i, 6).value
            next_date = Date(next_datetime.split()[0])
            #День изменился
            if cur_date.day != next_date.day or cur_date.month != next_date.month or cur_date.year != next_date.year:
                data_datetime = cur_date.GetDate() + '!' + enter_time + '!' + exit_time
                SaveDataToTabel(data_worker, data_datetime, date_end, count_all_sheets)
                cur_date = next_date
                cur_datetime = next_datetime
                enter_time = '00:00:00'
                exit_time = '24:00:00'
                #Приход
                if user_worksheet.cell(i, 4).value == 'КПП производства':
                    enter_time = cur_datetime.split()[1]
                #Уход
                else:
                    exit_time = cur_datetime.split()[1]
            else:
            #Это всё ещё текущий день
                #Сотрудник ушел
                if user_worksheet.cell(i, 4).value != 'КПП производства':
                    exit_time = next_datetime.split()[1]
                #Сотрудник пришел
                else:
                    exit_time = '24:00:00'
        data_datetime = cur_date.GetDate() + '!' + enter_time + '!' + exit_time
        SaveDataToTabel(data_worker, data_datetime, date_end, count_all_sheets)
    SetFormat()

def SetFormat():
    _thin = Side(border_style="medium", color="000000")
    workbook = openpyxl.load_workbook('Табель.xlsx')
    for sheet in workbook:
        for row in range(4, sheet.max_row + 1):
            for column in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, column)
                cell.border = Border(top=_thin, left=_thin, right=_thin, bottom=_thin)
                cell.font = Font(name="Times New Roman", size=12, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    workbook.save('Табель.xlsx')

def SaveDataToTabel(_data_worker, _data_datetime, _date_end, _count_of_all_sheets):
    global list_count_fio_in_sheets
    workbook = openpyxl.load_workbook('Табель.xlsx')
    _fio, _user_position_name, _user_division_name = _data_worker.split('!')
    _cur_date, _enter_time, _end_time = _data_datetime.split('!')
    number_of_sheet = _count_of_all_sheets - NumberOfSheets(_cur_date, _date_end)
    worksheet = workbook[str(number_of_sheet)]
    #Заполнение ячеек
    #Номер в списке
    cur_num = list_count_fio_in_sheets[number_of_sheet]
    if cur_num != 0:
        if _fio != worksheet.cell(cur_num, 2).value:
            cur_num += 1
            list_count_fio_in_sheets[number_of_sheet] += 1
    else:
        cur_num = 4
        list_count_fio_in_sheets[number_of_sheet] = 4
    cell = worksheet.cell(cur_num, 1)
    cell.fill = PatternFill("solid", fgColor="EEEEEE")
    worksheet.cell(cur_num, 1).value = cur_num - 3
    #ФИО
    cell = worksheet.cell(cur_num, 2)
    cell.fill = PatternFill("solid", fgColor="EEEEEE")
    worksheet.cell(cur_num, 2).value = _fio
    #Должность
    cell = worksheet.cell(cur_num, 3)
    cell.fill = PatternFill("solid", fgColor="EEEEEE")
    worksheet.cell(cur_num, 3).value = _user_division_name
    #Отдел
    cell = worksheet.cell(cur_num, 4)
    cell.fill = PatternFill("solid", fgColor="EEEEEE")
    worksheet.cell(cur_num, 4).value = _user_position_name
    #Время табеля
    column_date = Date(_cur_date).day
    if column_date > 15:
        column_date -= 15
    column_date += 4
    worksheet.cell(cur_num, column_date).value = f'{_enter_time[:5]}\n{_end_time[:5]}'
    workbook.save('Табель.xlsx')
    return True

if __name__ == '__main__':
    #Загрузка параметров работы программы
    try:
        with open("config.json", 'r', encoding='ANSI') as file:
            config = json.load(file)
    except:
        with open("config.json", 'w', encoding='ANSI') as file:
            config = dict()
            config['serverIP'] = '192.168.9.52'
            config['login'] = 'admin'
            config['password'] = 'adminfarm123'
            config['dateBegin'] = '2024-06-01'
            config['dateEnd'] = '2024-12-31'
            config['wrong_words'] = ["Полный", "1", "кпп"]
            json.dump(config, file, indent=4)
    tk = GetToken(config)
    #При отсутствии токена, завершается программа
    if tk == '':
        print('Не получен токен авторизации. Программа завершена с ошибкой')
        os.kill(os.getpid(),9)
    accessReport = GetAccessReport(config, tk)
    arrayAccessReport = StringToArray(accessReport)
    print('----------------Выгрузка информации о проходах----------------')
    if not SaveReports(arrayAccessReport):
        print('Неверно заполнены таблицы о проходах сотрудников. Программа завершена с ошибкой')
        os.kill(os.getpid(), 9)
    print('----------------Загрузка главного табеля----------------')
    count_of_work_sheets = NumberOfSheets(config['dateBegin'], config['dateEnd'])
    CreateTabelTemplate(count_of_work_sheets, config['dateBegin'])
    list_count_fio_in_sheets = [0 for i in range(count_of_work_sheets)]
    LoadMainTabel(config['dateEnd'], count_of_work_sheets)
    print('----------------Завершение работы программы-----------------')
    #day = Date('2024.02.28')